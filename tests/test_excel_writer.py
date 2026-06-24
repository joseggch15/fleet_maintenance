# -*- coding: utf-8 -*-
import os
import datetime

import openpyxl
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.formula import ArrayFormula

import mapping
import source_reader
import excel_writer


def _open(path):
    return openpyxl.load_workbook(path)[excel_writer.SHEET_NAME]


def test_append_writes_rows_after_last_data(target_workbook, source_workbook):
    subs = source_reader.read_submissions(source_workbook)
    rows = mapping.submissions_to_rows(subs)
    result = excel_writer.append_rows(target_workbook, rows,
                                      make_backup_copy=False)
    assert result["written"] == 2
    assert result["first_row"] == 4      # tras filas 2 y 3 existentes
    assert result["last_row"] == 5

    ws = _open(target_workbook)
    assert ws["C4"].value == 836         # Vehicle ID 1a submission
    assert ws["B4"].value == datetime.datetime(2026, 5, 15)
    assert ws["N4"].value == "SMU"
    assert ws["G4"].value == 14599
    assert ws["C5"].value == 3141        # 2a submission (Truck Tag)
    assert ws["N5"].value == "TAG"
    assert ws["G5"].value == 9176        # hourmeter


def test_formula_columns_reproduced(target_workbook, source_workbook):
    subs = source_reader.read_submissions(source_workbook)
    rows = mapping.submissions_to_rows(subs)
    excel_writer.append_rows(target_workbook, rows, make_backup_copy=False)

    ws = _open(target_workbook)
    # Columna A: misma formula estructurada que la plantilla.
    assert ws["A4"].value == mapping.A_FORMULA
    # Columna D: ArrayFormula con ref propio por fila.
    d4 = ws["D4"].value
    assert isinstance(d4, ArrayFormula)
    assert d4.ref == "D4"
    assert "Table33" in d4.text


def test_table_ref_extended(target_workbook, source_workbook):
    subs = source_reader.read_submissions(source_workbook)
    rows = mapping.submissions_to_rows(subs)
    excel_writer.append_rows(target_workbook, rows, make_backup_copy=False)

    wb = openpyxl.load_workbook(target_workbook)
    ws = wb[excel_writer.SHEET_NAME]
    tbl = dict(ws.tables)["Table3"]
    _, _, _, max_row = range_boundaries(tbl.ref)
    assert max_row == 5


def test_number_format_preserved(target_workbook, source_workbook):
    subs = source_reader.read_submissions(source_workbook)
    rows = mapping.submissions_to_rows(subs)
    excel_writer.append_rows(target_workbook, rows, make_backup_copy=False)
    ws = _open(target_workbook)
    assert ws["G4"].number_format == "0.0"
    assert ws["B4"].number_format == "[$-409]mmmm\\ d\\,\\ yyyy;@"


def test_backup_created(target_workbook, source_workbook):
    subs = source_reader.read_submissions(source_workbook)
    rows = mapping.submissions_to_rows(subs)
    result = excel_writer.append_rows(target_workbook, rows,
                                      make_backup_copy=True)
    assert result["backup"] and os.path.isfile(result["backup"])


def test_empty_rows_noop(target_workbook):
    result = excel_writer.append_rows(target_workbook, [],
                                      make_backup_copy=False)
    assert result["written"] == 0
    ws = _open(target_workbook)
    # Sin filas nuevas: la ultima fila de datos sigue siendo la 3.
    assert ws["C4"].value is None
