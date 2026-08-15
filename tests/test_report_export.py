# -*- coding: utf-8 -*-
"""Excel exportado: hojas, contenido, graficas nativas e idioma."""
import datetime
import os

import openpyxl
import pytest

import analytics
import i18n
import report_export
import tag_reader


@pytest.fixture(autouse=True)
def spanish():
    """Cada prueba arranca en espanol salvo que cambie el idioma a proposito."""
    i18n.set_language(i18n.ES)
    yield
    i18n.set_language(i18n.ES)


def _inspections():
    return [
        {"date": "2025-01-10", "vehicle_id": "829", "fms_id": "829",
         "system_fitted": "Y", "equipment_hours": "31093.3", "fms_hours": "3509",
         "status": "VIU OK", "inlets": "1", "addl_inlets_locked": "N",
         "drain_valves_locked": "N", "fast_fill_leaking": "N",
         "smu_tags": "SMU", "equipment_type": "HAUL TRUCK", "remarks": "",
         "inspectors": "RANCHO", "owner": "NEWMONT", "remedial": "",
         "source_file": "export.xlsx"},
        {"date": "2025-02-03", "vehicle_id": "802", "fms_id": "802",
         "system_fitted": "Y", "equipment_hours": "45814.3", "fms_hours": "",
         "status": "NO VIU", "inlets": "1", "addl_inlets_locked": "N",
         "drain_valves_locked": "N", "fast_fill_leaking": "N",
         "smu_tags": "TAG", "equipment_type": "LIGHT VEHICLE", "remarks": "",
         "inspectors": "REGILLIO", "owner": "SEMC", "remedial": "",
         "source_file": "export.xlsx"},
    ]


def _movements():
    return [
        {"move_type": tag_reader.MOVE_NEW, "date": "2026-06-01",
         "equipment_id": "LVE2136", "tag": "CA:CC:A3:FA:4C:2B",
         "device_type": "SMU", "cost_center": "14387",
         "department": "MINE_OPS", "product": "Diesel", "changed_by": "",
         "type_inferred": 0, "source_file": "semana.xlsx", "note": ""},
        {"move_type": tag_reader.MOVE_REMOVAL, "date": "2026-06-02",
         "equipment_id": "R-1480", "tag": "56B2D9A6", "device_type": "TAG",
         "cost_center": "14387", "department": "SUSTAINING CAPEX",
         "product": "Diesel", "changed_by": "", "type_inferred": 1,
         "source_file": "semana.xlsx",
         "note": "%s:19/19/2025" % tag_reader.NOTE_BAD_DATE},
    ]


# ---------------------------------------------------------------------------
# Reporte de mantenimiento
# ---------------------------------------------------------------------------
def test_maintenance_report_has_the_three_sheets(tmp_path):
    path = str(tmp_path / "reporte.xlsx")
    report_export.export_maintenance(path, _inspections(), {"2025-01": 400})
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == [i18n.t("sheet.fulllist"), i18n.t("sheet.pivot"),
                             i18n.t("sheet.notes")]


def test_full_list_writes_dates_and_numbers_typed(tmp_path):
    path = str(tmp_path / "reporte.xlsx")
    report_export.export_maintenance(path, _inspections(), {})
    ws = openpyxl.load_workbook(path)[i18n.t("sheet.fulllist")]
    header = next(r for r in range(1, 10)
                  if ws.cell(row=r, column=2).value == i18n.t("col.date_target"))
    assert ws.cell(row=header + 1, column=2).value == datetime.datetime(2025, 1, 10)
    # Las horas deben ir como numero para poder ordenarse y sumarse.
    assert ws.cell(row=header + 1, column=6).value == 31093.3
    # Columna A: el mes en texto, como en el maestro.
    assert ws.cell(row=header + 1, column=1).value == "ene-25"


def test_pivot_sheet_carries_both_charts(tmp_path):
    path = str(tmp_path / "reporte.xlsx")
    report_export.export_maintenance(path, _inspections(), {"2025-01": 400})
    ws = openpyxl.load_workbook(path)[i18n.t("sheet.pivot")]
    kinds = sorted(type(chart).__name__ for chart in ws._charts)
    # Las barras y la linea de % viajan en un mismo grafico combinado.
    assert kinds == ["BarChart", "PieChart"]


def test_pivot_counts_match_the_analytics(tmp_path):
    rows = _inspections()
    path = str(tmp_path / "reporte.xlsx")
    report_export.export_maintenance(path, rows, {"2025-01": 400})
    ws = openpyxl.load_workbook(path)[i18n.t("sheet.pivot")]
    pivot = analytics.build_pivot(rows)
    values = {ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)}
    for vehicle in pivot.vehicles:
        assert vehicle in values


def test_export_follows_the_interface_language(tmp_path):
    i18n.set_language(i18n.EN)
    path = str(tmp_path / "report.xlsx")
    report_export.export_maintenance(path, _inspections(), {})
    wb = openpyxl.load_workbook(path)
    assert "Full List 2024-2025" in wb.sheetnames
    assert wb[wb.sheetnames[0]].cell(row=1, column=1).value.startswith(
        "Fleet maintenance")


def test_window_limits_the_pivot_but_not_the_full_list(tmp_path):
    path = str(tmp_path / "reporte.xlsx")
    report_export.export_maintenance(path, _inspections(), {}, ["2025-02"])
    wb = openpyxl.load_workbook(path)
    full = wb[i18n.t("sheet.fulllist")]
    pivot = wb[i18n.t("sheet.pivot")]
    labels = {pivot.cell(row=r, column=1).value
              for r in range(1, pivot.max_row + 1)}
    assert "802" in labels and "829" not in labels
    # La hoja de inspecciones conserva las dos filas.
    vehicles = {full.cell(row=r, column=3).value
                for r in range(1, full.max_row + 1)}
    assert {"829", "802"} <= vehicles


def test_report_with_no_rows_still_opens(tmp_path):
    path = str(tmp_path / "vacio.xlsx")
    report_export.export_maintenance(path, [], {})
    assert openpyxl.load_workbook(path).sheetnames


# ---------------------------------------------------------------------------
# Consolidado de tags
# ---------------------------------------------------------------------------
def test_tag_export_sheets_and_chart(tmp_path):
    path = str(tmp_path / "tags.xlsx")
    report_export.export_tags(path, _movements())
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == [i18n.t("sheet.taginstalled"),
                            i18n.t("sheet.tagsummary"), i18n.t("sheet.notes")]
    assert len(wb[i18n.t("sheet.tagsummary")]._charts) == 1


def test_tag_export_resolves_year_month_and_week(tmp_path):
    path = str(tmp_path / "tags.xlsx")
    report_export.export_tags(path, _movements())
    ws = openpyxl.load_workbook(path)[i18n.t("sheet.taginstalled")]
    header = next(r for r in range(1, 10)
                  if ws.cell(row=r, column=1).value == i18n.t("col.num"))
    assert ws.cell(row=header + 1, column=11).value == 2026      # ano
    assert ws.cell(row=header + 1, column=12).value == 6         # mes
    assert ws.cell(row=header + 1, column=13).value == \
        datetime.datetime(2026, 6, 1)                            # lunes


def test_tag_export_translates_movements_and_notes(tmp_path):
    path = str(tmp_path / "tags.xlsx")
    report_export.export_tags(path, _movements())
    ws = openpyxl.load_workbook(path)[i18n.t("sheet.taginstalled")]
    texts = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
    assert i18n.tr_value(tag_reader.MOVE_REMOVAL) in texts
    notes = [ws.cell(row=r, column=16).value for r in range(1, ws.max_row + 1)]
    assert any(note and "19/19/2025" in note for note in notes)


def test_summary_counts_installed_without_removals(tmp_path):
    path = str(tmp_path / "tags.xlsx")
    report_export.export_tags(path, _movements())
    ws = openpyxl.load_workbook(path)[i18n.t("sheet.tagsummary")]
    header = next(r for r in range(1, 10)
                  if ws.cell(row=r, column=1).value == i18n.t("col.year"))
    assert ws.cell(row=header + 1, column=3).value == 1     # SMU
    assert ws.cell(row=header + 1, column=4).value == 0     # TAG (el otro fue retiro)
    assert ws.cell(row=header + 1, column=6).value == 1     # retiros


def test_tag_export_groups_by_the_requested_grain(tmp_path):
    """El resumen cambia de cubeta; el detalle de movimientos no."""
    for grain in analytics.GRAINS:
        path = str(tmp_path / ("tags_%s.xlsx" % grain))
        report_export.export_tags(path, _movements(), grain)
        wb = openpyxl.load_workbook(path)
        summary = wb[i18n.t("sheet.tagsummary")]
        data = wb[i18n.t("sheet.taginstalled")]
        rows = len(analytics.tag_by_period(_movements(), grain))
        # Titulo + subtitulo + blanco + encabezado + una fila por cubeta.
        assert summary.max_row == 4 + rows
        assert len(summary._charts) == 1
        assert data.max_row == 6      # el detalle no depende del grano


def test_monthly_grain_keeps_the_year_month_pair(tmp_path):
    path = str(tmp_path / "tags.xlsx")
    report_export.export_tags(path, _movements(), analytics.GRAIN_MONTH)
    ws = openpyxl.load_workbook(path)[i18n.t("sheet.tagsummary")]
    header = next(r for r in range(1, 10)
                  if ws.cell(row=r, column=1).value == i18n.t("col.year"))
    assert ws.cell(row=header, column=2).value == i18n.t("col.month")


def test_other_grains_use_a_single_period_column(tmp_path):
    path = str(tmp_path / "tags.xlsx")
    report_export.export_tags(path, _movements(), analytics.GRAIN_DAY)
    ws = openpyxl.load_workbook(path)[i18n.t("sheet.tagsummary")]
    header = next(r for r in range(1, 10)
                  if ws.cell(row=r, column=1).value == i18n.t("grain.col_day"))
    assert ws.cell(row=header, column=2).value == "SMU"
    # El dia va como fecha real, para poder ordenarlo y agruparlo en Excel.
    assert ws.cell(row=header + 1, column=1).value == \
        datetime.datetime(2026, 6, 1)


def test_unknown_grain_falls_back_to_monthly(tmp_path):
    path = str(tmp_path / "tags.xlsx")
    report_export.export_tags(path, _movements(), "quincenal")
    ws = openpyxl.load_workbook(path)[i18n.t("sheet.tagsummary")]
    assert ws.cell(row=4, column=1).value == i18n.t("col.year")


def test_suggested_name_changes_with_the_language(tmp_path):
    assert report_export.suggested_name("tags").startswith("Tags Instalados")
    i18n.set_language(i18n.EN)
    assert report_export.suggested_name("tags").startswith("Tag Installed")
    assert report_export.default_path(str(tmp_path), "report").endswith(".xlsx")
    assert os.path.dirname(
        report_export.default_path(str(tmp_path), "report")) == str(tmp_path)
