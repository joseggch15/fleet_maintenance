# -*- coding: utf-8 -*-
"""Fixtures sinteticas para las pruebas (no usan los archivos reales del cliente)."""
import os
import sys
import datetime

import pytest
import openpyxl
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.formula import ArrayFormula

# Permite importar mapping/source_reader/excel_writer desde la raiz del repo.
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import mapping  # noqa: E402


@pytest.fixture
def target_workbook(tmp_path):
    """Crea un Excel destino minimo que imita 'Full List 2024-2025':

    - encabezados A..S en la fila 1,
    - dos filas de datos (2 y 3) con la formula de A, una fecha en B, vehiculo
      en C y la formula de matriz en D,
    - tabla Table3 sobre A1:S3,
    - columna G con formato de numero '0.0' para verificar que se conserva.
    """
    path = os.path.join(str(tmp_path), "target.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = excel_sheet_name()

    cols = list(mapping.TARGET_HEADERS.keys())  # A..S
    for c, letter in enumerate(cols, start=1):
        ws.cell(row=1, column=c, value=mapping.TARGET_HEADERS[letter])

    def fill_data_row(r, date, vehicle, fms):
        ws["A%d" % r] = mapping.A_FORMULA
        ws["B%d" % r] = date
        ws["B%d" % r].number_format = "[$-409]mmmm\\ d\\,\\ yyyy;@"
        ws["C%d" % r] = vehicle
        ws["D%d" % r] = ArrayFormula(ref="D%d" % r, text=mapping.D_FORMULA_TEXT)
        ws["E%d" % r] = fms
        ws["F%d" % r] = "Y"
        ws["G%d" % r] = 100.0
        ws["G%d" % r].number_format = "0.0"
        ws["I%d" % r] = "VIU OK"

    fill_data_row(2, datetime.datetime(2024, 1, 3), 829, 829)
    fill_data_row(3, datetime.datetime(2024, 1, 6), 802, 802)

    ws.add_table(Table(displayName="Table3", ref="A1:S3"))
    wb.save(path)
    return path


@pytest.fixture
def source_workbook(tmp_path):
    """Crea un Excel de submissions que imita el export del formulario."""
    path = os.path.join(str(tmp_path), "source.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Form Submissions"
    ws["A1"] = "Report: ... Submissions Report"
    ws["A2"] = "Generated: 2026-06-24 01:51:00"

    headers = [
        mapping.H_ID, mapping.H_CODE, "Submitted By", mapping.H_SUBMITTED,
        "Form Title", "Form Description", "Form Is Public",
        "Form Creator Username", "Form Creator Email",
        "Form Creator Environment Name", "Created At", "Updated At",
        mapping.H_REVISION, mapping.H_VEHICLE, mapping.H_FMS, mapping.H_TAGTYPE,
        mapping.H_KM, mapping.H_HOUR, mapping.H_EQUIP, mapping.H_INLETS,
        mapping.H_TAGS, mapping.H_ADDLOCK, mapping.H_DRAIN, mapping.H_COMPANY,
        mapping.H_INSPECTOR, mapping.H_REMARKS, mapping.H_SIGNATURE,
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=h)

    rows = [
        # SMU con hourmeter vacio -> usa kilometer
        {mapping.H_ID: 76, mapping.H_CODE: "MOFFM-1",
         mapping.H_REVISION: "15/05/2026 13:39:00", mapping.H_VEHICLE: 836,
         mapping.H_FMS: 836, mapping.H_TAGTYPE: "SMU", mapping.H_KM: 14599,
         mapping.H_EQUIP: "Water Truck", mapping.H_INLETS: 1,
         mapping.H_TAGS: 1, mapping.H_ADDLOCK: "Not applicable",
         mapping.H_DRAIN: "Not applicable", mapping.H_COMPANY: "Newmont",
         mapping.H_INSPECTOR: "Rouchon Radi", mapping.H_REMARKS: "via ok"},
        # Truck Tag con hourmeter -> usa hourmeter; locked Yes -> Y
        {mapping.H_ID: 100, mapping.H_CODE: "MOFFM-4",
         mapping.H_REVISION: "17/05/2026 14:35:00", mapping.H_VEHICLE: 3141,
         mapping.H_FMS: 3141, mapping.H_TAGTYPE: "Truck Tag",
         mapping.H_HOUR: 9176, mapping.H_KM: 5000,
         mapping.H_EQUIP: "Light Trucks", mapping.H_INLETS: 1,
         mapping.H_TAGS: 1, mapping.H_ADDLOCK: "Yes",
         mapping.H_DRAIN: "Not applicable", mapping.H_COMPANY: "Semc",
         mapping.H_INSPECTOR: "Rouchon Radi"},
    ]
    for i, rec in enumerate(rows, start=5):
        for c, h in enumerate(headers, start=1):
            if h in rec:
                ws.cell(row=i, column=c, value=rec[h])
    wb.save(path)
    return path


@pytest.fixture
def weekly_tag_files(tmp_path):
    """Carpeta que imita 'Tag Installed Per Week' con sus tres formatos.

    - un archivo viejo sin columna TYPE y con los alias 'CC'/'Dept',
    - un archivo nuevo con TYPE y Product,
    - una subcarpeta con otro archivo nuevo que SOLAPA un dia con el anterior
      (misma fila repetida), para verificar la deduplicacion,
    - una hoja 'Summary' que no es data y debe ignorarse.
    """
    folder = tmp_path / "Tag Installed Per Week"
    folder.mkdir()
    sub = folder / "Inventory Tag Installed 03062026"
    sub.mkdir()

    old = openpyxl.Workbook()
    ws = old.active
    ws.title = "Sheet1"
    ws.append(["DATE", "ID", "Tag", "CC", "Dept"])
    ws.append([datetime.datetime(2025, 1, 10), "C-155", "56B2B853", 14386,
               "mine_ops"])
    ws.append(["19/19/2025", "CPR0988", "56B5609E", 14413, "PROCESS OPS"])
    old.save(str(folder / "Inventory Tag Installed 12012025.xlsx"))

    new = openpyxl.Workbook()
    ws = new.active
    ws.title = "Tag Installed"
    ws.append(["TYPE", "DATE", "ID", "Tag", "Cost Center", "Department",
               "Product"])
    ws.append(["NEW INSTALLATION", datetime.datetime(2026, 6, 1), "LVE2136",
               "CA:CC:A3:FA:4C:2B", 14387, "MINE_OPS", "Diesel"])
    ws.append(["REMOVAL", datetime.datetime(2026, 6, 2), "R-1480",
               "56B2D9A6", 14387, "SUSTAINING CAPEX", "Diesel"])
    ws.append(["Tag updated ", datetime.datetime(2026, 6, 3), "C-201",
               "56B36ACF", 14386, "MINE_OPS", None])
    summary = new.create_sheet("Summary")
    summary.append(["KPI", "Value"])
    summary.append(["Total", 3])
    new.save(str(folder / "Inventory Tag Installed 01062026.xlsx"))

    overlap = openpyxl.Workbook()
    ws = overlap.active
    ws.title = "Tag Installed"
    ws.append(["TYPE", "DATE", "ID", "Tag", "Cost Center", "Department"])
    # Repetida del archivo anterior: no debe entrar dos veces.
    ws.append(["REMOVAL", datetime.datetime(2026, 6, 2), "R-1480",
               "56B2D9A6", 14387, "SUSTAINING CAPEX"])
    ws.append(["REPLACEMENT", datetime.datetime(2026, 6, 5), "LTR0371",
               "56B56D63", 15086, "EMULSION PLANT"])
    overlap.save(str(sub / "Inventory Tag Installed 03062026-10062026.xlsx"))
    return str(folder)


@pytest.fixture
def temp_store(tmp_path):
    """Base local vacia y aislada de la del usuario."""
    import store
    previous = store.path()
    store.set_path(os.path.join(str(tmp_path), "test.sqlite3"))
    store.init()
    yield store
    store.set_path(previous)


def excel_sheet_name():
    import excel_writer
    return excel_writer.SHEET_NAME
