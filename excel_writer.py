# -*- coding: utf-8 -*-
"""
Escribe filas nuevas en la hoja 'Full List 2024-2025' del Excel maestro
'Fleet Tag Inventory and Maintenance.xlsx', conservando su estructura:

  - copia estilos (fuente, relleno, bordes, alineacion, formato de numero)
    desde la ultima fila de datos como plantilla;
  - reproduce las formulas de las columnas A ('Date') y D ('Verified'),
    incluyendo la formula de matriz (ArrayFormula) de D con su nuevo `ref`;
  - extiende el rango de la tabla `Table3` para incluir las filas nuevas, de
    modo que Excel no reporte 'We found a problem with some content';
  - hace un respaldo con timestamp del archivo antes de modificarlo.

Las hojas de pivotes (PIVOT SUMMARY, INSTALLATION SUMMARY, COUNT NEWMONT) y
las hojas de inventario derivan de esta tabla: tras importar, basta abrir el
Excel y usar 'Datos > Actualizar todo' para que reflejen las filas nuevas.
"""
from __future__ import annotations

import copy
import datetime
import os
import shutil

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.formula import ArrayFormula

import mapping

SHEET_NAME = "Full List 2024-2025"
TABLE_NAME = "Table3"
_LAST_COL = column_index_from_string("S")   # 19, ultima columna de la tabla
_COL_DATE = column_index_from_string("B")   # 2
_COL_VEHICLE = column_index_from_string("C")  # 3


# ---------------------------------------------------------------------------
# Localizacion de la fila plantilla / ultima fila de datos
# ---------------------------------------------------------------------------
def _last_data_row(ws) -> int:
    """Ultima fila con datos reales (valor en col B 'Date' o C 'Vehicle ID').

    La hoja puede tener una fila final 'vacia' que solo trae las formulas de A
    y D (fila plantilla de la tabla); esa no cuenta como dato.
    """
    last = 1  # fila 1 = encabezado
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=_COL_DATE).value not in (None, "") or \
                ws.cell(row=r, column=_COL_VEHICLE).value not in (None, ""):
            last = r
    return last


# ---------------------------------------------------------------------------
# Escritura de una fila
# ---------------------------------------------------------------------------
def _copy_style(src, tgt) -> None:
    if src.has_style:
        tgt.font = copy.copy(src.font)
        tgt.fill = copy.copy(src.fill)
        tgt.border = copy.copy(src.border)
        tgt.alignment = copy.copy(src.alignment)
        tgt.number_format = src.number_format
        tgt.protection = copy.copy(src.protection)


def _apply_formula(src, tgt, target_row: int, letter: str) -> None:
    """Reproduce la formula de una columna calculada (A o D).

    Las formulas usan referencias estructuradas ([#This Row]), por lo que el
    texto es identico en cada fila; solo la ArrayFormula necesita un `ref`
    propio por fila.
    """
    val = src.value
    if isinstance(val, ArrayFormula):
        tgt.value = ArrayFormula(ref="%s%d" % (letter, target_row),
                                 text=val.text)
    elif isinstance(val, str) and val.startswith("="):
        tgt.value = val
    elif letter == "A":
        tgt.value = mapping.A_FORMULA
    elif letter == "D":
        tgt.value = ArrayFormula(ref="D%d" % target_row,
                                 text=mapping.D_FORMULA_TEXT)


def _write_row(ws, target_row: int, template_row: int, rowdata: dict) -> None:
    for col in range(1, _LAST_COL + 1):
        letter = get_column_letter(col)
        src = ws.cell(row=template_row, column=col)
        tgt = ws.cell(row=target_row, column=col)
        _copy_style(src, tgt)
        if letter in mapping.FORMULA_COLUMNS:
            _apply_formula(src, tgt, target_row, letter)
        elif letter in rowdata:
            tgt.value = rowdata[letter]


def _extend_table(ws, last_row: int) -> None:
    try:
        tables = dict(ws.tables)
    except AttributeError:
        return
    table = tables.get(TABLE_NAME)
    if table is None and tables:
        table = next(iter(tables.values()))
    if table is None:
        return
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    if last_row > max_row:
        table.ref = "%s%d:%s%d" % (
            get_column_letter(min_col), min_row,
            get_column_letter(max_col), last_row)


def make_backup(path: str) -> str:
    base, ext = os.path.splitext(path)
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = "%s.backup_%s%s" % (base, stamp, ext)
    shutil.copy2(path, backup)
    return backup


# ---------------------------------------------------------------------------
# API publica
# ---------------------------------------------------------------------------
def append_rows(target_path: str, rows: list, make_backup_copy: bool = True,
                progress_cb=None) -> dict:
    """Agrega `rows` (lista de dicts {col: valor}) a 'Full List 2024-2025'.

    Retorna un dict con: written, first_row, last_row, sheet, backup.
    Lanza ValueError si falta la hoja o no hay fila plantilla.
    """
    rows = [r for r in rows if r is not None]
    if not rows:
        return {"written": 0, "first_row": None, "last_row": None,
                "sheet": SHEET_NAME, "backup": None}

    backup = make_backup(target_path) if make_backup_copy else None

    wb = openpyxl.load_workbook(target_path)
    try:
        if SHEET_NAME not in wb.sheetnames:
            raise ValueError(
                "El Excel destino no tiene la hoja '%s'." % SHEET_NAME)
        ws = wb[SHEET_NAME]
        template_row = _last_data_row(ws)
        if template_row < 2:
            raise ValueError(
                "La hoja '%s' no tiene filas de datos para usar como "
                "plantilla de estilo/formula." % SHEET_NAME)

        start = template_row + 1
        total = len(rows)
        for i, rowdata in enumerate(rows):
            target_row = start + i
            if progress_cb:
                try:
                    progress_cb(i + 1, total,
                                "Escribiendo fila %d" % target_row)
                except Exception:
                    pass
            _write_row(ws, target_row, template_row, rowdata)

        last_row = start + total - 1
        _extend_table(ws, last_row)
        wb.save(target_path)
    finally:
        wb.close()

    return {"written": total, "first_row": start, "last_row": last_row,
            "sheet": SHEET_NAME, "backup": backup}
