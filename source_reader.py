# -*- coding: utf-8 -*-
"""
Lectura del Excel exportado del formulario:
'merian_ops - Form - Fleet maintenance - Submissions'.

La hoja 'Form Submissions' tiene un par de filas de titulo y el encabezado
real en la fila donde aparece la columna 'Id' (tipicamente fila 4); los datos
empiezan en la fila siguiente. Este modulo encuentra el encabezado de forma
robusta (por nombre, no por posicion) y devuelve una lista de submissions,
cada una como un dict {encabezado: valor}.
"""
from __future__ import annotations

import openpyxl

import mapping

SHEET_NAME = "Form Submissions"
_HEADER_KEY = mapping.H_ID            # "Id"
_MAX_HEADER_SCAN = 20                 # filas a inspeccionar buscando el header


def _clean(value):
    return value.strip() if isinstance(value, str) else value


def _find_header(ws):
    """Devuelve (fila_header, {indice_columna: encabezado})."""
    for r, row in enumerate(
            ws.iter_rows(min_row=1, max_row=_MAX_HEADER_SCAN,
                         values_only=True), start=1):
        for idx, val in enumerate(row):
            if isinstance(val, str) and val.strip() == _HEADER_KEY:
                headers = {i: _clean(v)
                           for i, v in enumerate(row) if v not in (None, "")}
                return r, headers
    raise ValueError(
        "No se encontro la fila de encabezados (columna '%s') en la hoja '%s'."
        % (_HEADER_KEY, ws.title))


def read_submissions(path: str, sheet_name: str = SHEET_NAME) -> list:
    """Lee las submissions del formulario.

    Retorna una lista de dicts {encabezado: valor}. Se ignoran filas vacias y
    filas sin 'Id' ni 'Vehicle Id' (titulos, separadores, totales, etc.).
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
        header_row, headers = _find_header(ws)
        subs = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if all(v in (None, "") for v in row):
                continue
            record = {}
            for idx, name in headers.items():
                record[name] = row[idx] if idx < len(row) else None
            if record.get(mapping.H_ID) in (None, "") and \
                    record.get(mapping.H_VEHICLE) in (None, ""):
                continue
            subs.append(record)
        return subs
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Lectura de la hoja 'Full List 2024-2025' del maestro
# ---------------------------------------------------------------------------
FULL_LIST_SHEET = "Full List 2024-2025"


def _full_list_columns(header_row) -> dict:
    """{letra_de_columna: indice} para las columnas de datos de la hoja.

    Se buscan por encabezado y, si algun encabezado no coincide (el maestro
    tiene erratas historicas como 'UPDPATES' que alguien puede corregir), se
    cae a la posicion fija de la hoja. Las dos vias dan lo mismo mientras nadie
    reordene columnas; la busqueda por nombre cubre el caso de que alguien lo
    haga.
    """
    labels = {}
    for idx, value in enumerate(header_row):
        if isinstance(value, str) and value.strip():
            labels[" ".join(value.split()).upper()] = idx

    columns = {}
    for letter in mapping.DATA_COLUMNS:
        wanted = " ".join(mapping.TARGET_HEADERS[letter].split()).upper()
        if wanted in labels:
            columns[letter] = labels[wanted]
        else:
            columns[letter] = ord(letter) - ord("A")
    return columns


def read_full_list(path: str, sheet_name: str = FULL_LIST_SHEET,
                   progress_cb=None) -> list:
    """Lee el historico ya cargado en el maestro.

    Devuelve filas {columna: valor} con la misma forma que produce `mapping`,
    para que la base local pueda arrancar con los anos que ya estan en el Excel
    y no solo con las submissions nuevas. Sin esto el tablero mostraria los
    meses del ultimo export y los indicadores no coincidirian con los del
    maestro.

    Se saltan las filas sin fecha NI equipo: la tabla `Table3` termina con una
    fila plantilla que solo trae las formulas de 'Date' y 'Verified'.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                "El Excel no tiene la hoja '%s'." % sheet_name)
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            return []
        columns = _full_list_columns(header)

        total = max(ws.max_row - 1, 1)
        out = []
        for i, raw in enumerate(rows):
            if raw is None or all(v in (None, "") for v in raw):
                continue
            record = {}
            for letter, idx in columns.items():
                record[letter] = raw[idx] if idx < len(raw) else None
            if record.get("B") in (None, "") and record.get("C") in (None, ""):
                continue
            out.append(record)
            if progress_cb and i % 250 == 0:
                progress_cb(i + 1, total, "")
        return out
    finally:
        wb.close()


def submission_label(sub: dict) -> str:
    """Etiqueta corta para mostrar en logs/UI: 'MOFFM-12 (vehiculo 836)'."""
    code = sub.get(mapping.H_CODE) or sub.get(mapping.H_ID) or "?"
    vehicle = sub.get(mapping.H_VEHICLE)
    return "%s (vehiculo %s)" % (code, vehicle) if vehicle else str(code)
