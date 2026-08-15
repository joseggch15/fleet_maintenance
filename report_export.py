# -*- coding: utf-8 -*-
"""
Genera los Excel que el software exporta desde su base local.

Dos reportes, cada uno con las mismas hojas y graficas que el archivo del
cliente al que reemplaza:

  `export_maintenance`  hoja de inspecciones con el formato de
                        'Full List 2024-2025', hoja de resumen dinamico
                        (equipos x meses) con el bloque de indicadores y las
                        DOS graficas del maestro — la torta de revisados y las
                        barras con la linea de % de inspeccion — y una hoja de
                        notas.

  `export_tags`         consolidado de 'Tag Installed Per Week': la tabla de
                        movimientos unificada y el resumen de instalacion por
                        mes con su grafica de barras.

Las graficas son NATIVAS de Excel (openpyxl), no imagenes: el que reciba el
archivo puede filtrar la tabla, cambiar el rango y ver la grafica moverse. Por
eso los indicadores se escriben como celdas y no como texto.

Idioma: todo el contenido —nombres de hoja, encabezados, titulos de grafica y
notas— sale de `i18n`, asi que el archivo queda en el idioma que tenga la
ventana al exportar. Los identificadores de campo (Vehicle ID, FMS ID, Tag,
Cost Center) no se traducen nunca: son la jerga con la que se cruza contra el
FMS y contra los inventarios de Newmont.

Color: el Excel exportado se imprime y se comparte, asi que siempre sale con
la paleta CLARA aunque la ventana este en tema oscuro (`theme.excel_color`).
"""
from __future__ import annotations

import datetime
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import analytics
import i18n
import store
import tag_reader
import theme

_PCT_FORMAT = "0%"
_DATE_FORMAT = "DD/MM/YYYY"
# Meses que entran en la torta del resumen (ver `_add_summary_charts`).
_PIE_MONTHS = 12

_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ---------------------------------------------------------------------------
# Ayudas de formato
# ---------------------------------------------------------------------------
def _primary() -> str:
    return theme.excel_color("primary")


def _title(ws, row: int, text: str, subtitle: str = "") -> int:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=14, color=_primary())
    row += 1
    if subtitle:
        sub = ws.cell(row=row, column=1, value=subtitle)
        sub.font = Font(size=9, italic=True,
                        color=theme.excel_color("text_muted"))
        row += 1
    return row + 1


def _headers(ws, row: int, labels: list, widths: list = None) -> None:
    fill = PatternFill("solid", fgColor=_primary())
    for idx, label in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=idx, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = _BORDER
    ws.row_dimensions[row].height = 30
    for idx, width in enumerate(widths or [], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _as_table(ws, name: str, header_row: int, last_row: int,
              last_col: int) -> None:
    """Convierte el rango en tabla de Excel (filtro y bandas incluidos).

    Excel rechaza el archivo si dos tablas comparten nombre o si el nombre
    tiene espacios, de ahi la normalizacion.
    """
    if last_row <= header_row:
        return
    ref = "A%d:%s%d" % (header_row, get_column_letter(last_col), last_row)
    table = Table(displayName=name.replace(" ", "_"), ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)


def _note(ws, row: int, text: str) -> int:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(size=9, color=theme.excel_color("text_muted"))
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    return row + 1


def _stamp() -> str:
    return datetime.datetime.now().strftime("%d/%m/%Y %H:%M")


def _text_categories(chart, ref: Reference) -> None:
    """Asigna categorias de TEXTO a las series de un grafico.

    `chart.set_categories()` de openpyxl siempre escribe la referencia como
    numerica. Las etiquetas de mes ('ene-25', 'Jan-25') son texto, y con una
    referencia numerica Excel guarda una cache vacia: el eje se ve bien al
    abrir el archivo, pero cualquier visor que lea la cache en vez de las
    celdas muestra 1, 2, 3... en lugar de los meses.
    """
    chart.set_categories(ref)
    for series in chart.series:
        series.cat = AxDataSource(strRef=StrRef(f=ref))


def _sheet_notes(wb, lines: list) -> None:
    ws = wb.create_sheet(i18n.t("sheet.notes"))
    ws.column_dimensions["A"].width = 118
    row = _title(ws, 1, i18n.t("xls.notes_title"),
                 i18n.t("xls.generated", when=_stamp()))
    for line in lines:
        row = _note(ws, row, line)
        row += 0
    ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------------------
# Reporte de mantenimiento
# ---------------------------------------------------------------------------
_FULL_COLUMNS = (
    ("col.date", None, 10),                    # mmm-aa, calculada
    ("col.date_target", "date", 13),
    ("col.vehicle", "vehicle_id", 12),
    ("col.fms", "fms_id", 12),
    ("col.fitted", "system_fitted", 14),
    ("col.hours", "equipment_hours", 13),
    ("col.fms_hours", "fms_hours", 11),
    ("col.status", "status", 12),
    ("col.inlets", "inlets", 9),
    ("col.addl_locked", "addl_inlets_locked", 15),
    ("col.drain_locked", "drain_valves_locked", 15),
    ("col.leaking", "fast_fill_leaking", 15),
    ("col.smu_tags", "smu_tags", 11),
    ("col.equipment", "equipment_type", 20),
    ("col.remarks", "remarks", 34),
    ("col.inspectors", "inspectors", 18),
    ("col.owner", "owner", 14),
    ("col.remedial", "remedial", 28),
    ("col.source", "source_file", 26),
)

# Columnas que van al Excel como numero y no como texto. La base guarda todo
# como texto (es lo unico que garantiza que un ID como '0086' no pierda el
# cero), pero en el Excel las horas tienen que sumar y ordenarse como numeros.
_NUMERIC_FIELDS = ("equipment_hours", "fms_hours", "inlets")


def _sheet_full_list(wb, inspections: list) -> None:
    ws = wb.create_sheet(i18n.t("sheet.fulllist"))
    row = _title(ws, 1, i18n.t("xls.title_fulllist"),
                 "%s   ·   %s" % (i18n.t("xls.generated", when=_stamp()),
                                  i18n.t("xls.rows", n=len(inspections))))
    header_row = row
    _headers(ws, header_row, [i18n.t(key) for key, _f, _w in _FULL_COLUMNS],
             [w for _key, _f, w in _FULL_COLUMNS])

    for offset, record in enumerate(inspections, start=1):
        r = header_row + offset
        day = store.parse_date(record.get("date"))
        # Columna A del maestro: el mes en texto 'ene-25'. Alli es una formula
        # sobre la fecha; aqui se escribe ya resuelta porque el archivo se
        # genera desde cero y no hay tabla estructurada a la que referirse.
        ws.cell(row=r, column=1,
                value=i18n.month_label(day) if day else "")
        for idx, (_key, field, _w) in enumerate(_FULL_COLUMNS[1:], start=2):
            value = record.get(field)
            if field == "date":
                cell = ws.cell(row=r, column=idx, value=day)
                cell.number_format = _DATE_FORMAT
                continue
            if field in _NUMERIC_FIELDS:
                value = store.number(value)
            ws.cell(row=r, column=idx, value=value if value != "" else None)

    last_row = header_row + len(inspections)
    _as_table(ws, "Inspections", header_row, last_row, len(_FULL_COLUMNS))
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = ws.dimensions


def _write_kpi_block(ws, row: int, col: int, kpis: list, pivot,
                     fleet_pct) -> dict:
    """Escribe el bloque de indicadores y devuelve donde quedo cada fila.

    Es el mismo bloque del maestro (Z4:AS7): una columna por mes, una fila por
    indicador. Las graficas leen de aqui, no de la dinamica, porque lo que se
    grafica es el resumen y no las 800 filas de equipos.
    """
    header_row = row
    ws.cell(row=header_row, column=col,
            value=i18n.t("dash.table_month")).font = Font(
                bold=True, color=_primary())
    for i, kpi in enumerate(kpis):
        cell = ws.cell(row=header_row, column=col + 1 + i,
                       value=i18n.month_label(kpi.month))
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill("solid", fgColor="FFF2CC")
        cell.border = _BORDER

    rows = {
        "reviewed": (header_row + 1, "xls.kpi_reviewed"),
        "pct": (header_row + 2, "xls.kpi_pct"),
        "fleet": (header_row + 3, "xls.kpi_fleet"),
        "pct_fleet": (header_row + 4, "xls.kpi_pct_fleet"),
    }
    for _name, (r, key) in rows.items():
        label = ws.cell(row=r, column=col, value=i18n.t(key))
        label.font = Font(bold=True)
        label.border = _BORDER

    for i, kpi in enumerate(kpis):
        c = col + 1 + i
        cell = ws.cell(row=rows["reviewed"][0], column=c, value=kpi.reviewed)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _BORDER
        pct = ws.cell(row=rows["pct"][0], column=c,
                      value=kpi.pct if kpi.pct is not None else None)
        pct.number_format = _PCT_FORMAT
        pct.alignment = Alignment(horizontal="center")
        pct.border = _BORDER

    # Los dos totales del periodo ocupan una celda y se estiran bajo los meses,
    # igual que en el maestro.
    span = max(len(kpis), 1)
    total = ws.cell(row=rows["fleet"][0], column=col + 1,
                    value=pivot.maintained_fleet)
    total.alignment = Alignment(horizontal="center")
    pct_fleet = ws.cell(row=rows["pct_fleet"][0], column=col + 1,
                        value=fleet_pct)
    pct_fleet.number_format = _PCT_FORMAT
    pct_fleet.alignment = Alignment(horizontal="center")
    if span > 1:
        for r in (rows["fleet"][0], rows["pct_fleet"][0]):
            ws.merge_cells(start_row=r, start_column=col + 1,
                           end_row=r, end_column=col + span)

    ws.column_dimensions[get_column_letter(col)].width = 24
    return {"header": header_row,
            "reviewed": rows["reviewed"][0],
            "pct": rows["pct"][0],
            "first_col": col + 1,
            "last_col": col + span}


def _add_summary_charts(ws, block: dict, anchor_col: int) -> None:
    """La torta y las barras con linea, ancladas debajo del bloque de KPI."""
    cats = Reference(ws, min_col=block["first_col"], max_col=block["last_col"],
                     min_row=block["header"], max_row=block["header"])
    reviewed = Reference(ws, min_col=block["first_col"] - 1,
                         max_col=block["last_col"],
                         min_row=block["reviewed"], max_row=block["reviewed"])
    pct = Reference(ws, min_col=block["first_col"] - 1,
                    max_col=block["last_col"],
                    min_row=block["pct"], max_row=block["pct"])

    # La torta se recorta a los ultimos meses: con dos anos de columnas las
    # porciones quedan de 3% cada una y el grafico no dice nada. Las barras si
    # llevan el periodo completo, que es donde se ve la evolucion.
    pie_first = max(block["first_col"], block["last_col"] - _PIE_MONTHS + 1)
    pie_cats = Reference(ws, min_col=pie_first, max_col=block["last_col"],
                         min_row=block["header"], max_row=block["header"])
    pie_data = Reference(ws, min_col=pie_first, max_col=block["last_col"],
                         min_row=block["reviewed"], max_row=block["reviewed"])

    pie = PieChart()
    pie.title = i18n.t("xls.chart_pie")
    # `from_rows` porque el bloque es horizontal: un mes por columna.
    pie.add_data(pie_data, titles_from_data=False, from_rows=True)
    _text_categories(pie, pie_cats)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = True
    pie.height, pie.width = 9.5, 17.0
    ws.add_chart(pie, "%s%d" % (get_column_letter(anchor_col),
                                block["pct"] + 3))

    bar = BarChart()
    bar.type = "col"
    bar.title = i18n.t("xls.chart_bars")
    bar.add_data(reviewed, titles_from_data=True, from_rows=True)
    _text_categories(bar, cats)
    bar.y_axis.title = i18n.t("xls.axis_count")
    bar.x_axis.title = i18n.t("xls.axis_month")
    bar.dataLabels = DataLabelList()
    bar.dataLabels.showVal = True

    line = LineChart()
    line.add_data(pct, titles_from_data=True, from_rows=True)
    _text_categories(line, cats)
    # Eje secundario: el % va de 0 a 0,4 y el conteo de 0 a 250; sobre el mismo
    # eje la linea quedaria pegada al cero y no diria nada.
    line.y_axis.axId = 200
    line.y_axis.title = i18n.t("xls.axis_pct")
    line.y_axis.numFmt = _PCT_FORMAT
    line.y_axis.majorGridlines = None
    # 'max' manda el eje secundario al borde derecho. Con el valor por defecto
    # los dos ejes se dibujan encima uno del otro sobre el borde izquierdo.
    bar.y_axis.crosses = "max"
    bar += line
    bar.height, bar.width = 10.0, 26.0
    ws.add_chart(bar, "%s%d" % (get_column_letter(anchor_col),
                                block["pct"] + 24))


def _sheet_pivot(wb, pivot, kpis, fleet_pct) -> None:
    ws = wb.create_sheet(i18n.t("sheet.pivot"))
    years = sorted({m[:4] for m in pivot.months})
    span = years[0] if len(years) == 1 else (
        "%s-%s" % (years[0], years[-1]) if years else "")
    row = _title(ws, 1, i18n.t("xls.title_pivot", year=span),
                 i18n.t("xls.generated", when=_stamp()))

    header_row = row
    labels = [i18n.t("col.vehicle")] + \
             [i18n.month_label(m) for m in pivot.months] + \
             [i18n.t("col.grand_total")]
    _headers(ws, header_row, labels,
             [16] + [9] * len(pivot.months) + [12])

    for offset, vehicle in enumerate(pivot.vehicles, start=1):
        r = header_row + offset
        ws.cell(row=r, column=1, value=vehicle).font = Font(bold=True)
        for i, month in enumerate(pivot.months, start=2):
            count = pivot.count(vehicle, month)
            cell = ws.cell(row=r, column=i, value=count or None)
            cell.alignment = Alignment(horizontal="center")
        total = ws.cell(row=r, column=len(pivot.months) + 2,
                        value=pivot.row_total(vehicle))
        total.font = Font(bold=True)
        total.alignment = Alignment(horizontal="center")

    total_row = header_row + len(pivot.vehicles) + 1
    ws.cell(row=total_row, column=1,
            value=i18n.t("col.grand_total")).font = Font(bold=True)
    for i, month in enumerate(pivot.months, start=2):
        cell = ws.cell(row=total_row, column=i, value=pivot.col_total(month))
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    grand = ws.cell(row=total_row, column=len(pivot.months) + 2,
                    value=pivot.total_inspections)
    grand.font = Font(bold=True)
    grand.alignment = Alignment(horizontal="center")

    ws.freeze_panes = ws.cell(row=header_row + 1, column=2)

    kpi_col = len(pivot.months) + 4
    block = _write_kpi_block(ws, header_row, kpi_col, kpis, pivot, fleet_pct)
    _add_summary_charts(ws, block, kpi_col)


def export_maintenance(path: str, inspections: list, fleet_sizes: dict = None,
                       window: list = None, progress_cb=None) -> str:
    """Escribe el reporte de mantenimiento y devuelve la ruta.

    `window` (claves 'AAAA-MM') acota la hoja de resumen dinamico, igual que en
    el maestro, donde hay una hoja por periodo. La hoja de inspecciones lleva
    SIEMPRE todas las filas recibidas: es el respaldo del dato, y recortarla
    dejaria fuera el historico que el resumen no muestra.
    """
    inspections = list(inspections or [])
    pivot = analytics.build_pivot(inspections, window)
    kpis = analytics.monthly_kpis(pivot, fleet_sizes)
    fleet_pct = analytics.fleet_maintenance_pct(pivot, kpis)

    wb = Workbook()
    wb.remove(wb.active)
    if progress_cb:
        progress_cb(1, 3, i18n.t("sheet.fulllist"))
    _sheet_full_list(wb, inspections)
    if progress_cb:
        progress_cb(2, 3, i18n.t("sheet.pivot"))
    _sheet_pivot(wb, pivot, kpis, fleet_pct)
    if progress_cb:
        progress_cb(3, 3, i18n.t("sheet.notes"))

    notes = [i18n.t("note.fulllist"), i18n.t("note.pivot"), i18n.t("note.pct")]
    if pivot.undated:
        notes.append(i18n.t("note.undated", n=pivot.undated))
    notes.append(i18n.t("note.no_verified"))
    _sheet_notes(wb, notes)

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Consolidado de tags instalados
# ---------------------------------------------------------------------------
_TAG_COLUMNS = (
    ("col.num", None, 6),
    ("col.move_type", "move_type", 18),
    ("col.date", "date", 12),
    ("col.equipment_id", "equipment_id", 15),
    ("col.tag", "tag", 24),
    ("col.device", "device_type", 13),
    ("col.cost_center", "cost_center", 22),
    ("col.department", "department", 22),
    ("col.product", "product", 11),
    ("col.changed_by", "changed_by", 14),
    ("col.year", None, 8),
    ("col.month", None, 8),
    ("col.week", None, 13),
    ("col.source", "source_file", 34),
    ("col.inferred", "type_inferred", 12),
    ("col.note", "note", 34),
)


def _sheet_tag_data(wb, movements: list) -> None:
    ws = wb.create_sheet(i18n.t("sheet.taginstalled"))
    row = _title(ws, 1, i18n.t("xls.title_taginstalled"),
                 "%s   ·   %s" % (i18n.t("xls.generated", when=_stamp()),
                                  i18n.t("xls.rows", n=len(movements))))
    header_row = row
    _headers(ws, header_row, [i18n.t(key) for key, _f, _w in _TAG_COLUMNS],
             [w for _key, _f, w in _TAG_COLUMNS])

    for offset, record in enumerate(movements, start=1):
        r = header_row + offset
        day = store.parse_date(record.get("date"))
        monday = tag_reader.week_monday(day) if day else None
        ws.cell(row=r, column=1, value=offset)
        for idx, (_key, field, _w) in enumerate(_TAG_COLUMNS[1:], start=2):
            if field == "date":
                cell = ws.cell(row=r, column=idx, value=day)
                cell.number_format = _DATE_FORMAT
            elif field == "move_type":
                ws.cell(row=r, column=idx,
                        value=i18n.tr_value(record.get("move_type")))
            elif field == "type_inferred":
                ws.cell(row=r, column=idx,
                        value=i18n.tr_value("Y" if record.get("type_inferred")
                                            else "N"))
            elif field == "note":
                ws.cell(row=r, column=idx,
                        value=i18n.tr_note(record.get("note")) or None)
            elif field is None:
                pass
            else:
                value = record.get(field)
                ws.cell(row=r, column=idx, value=value if value != "" else None)
        # Ano, mes y semana se escriben resueltos: son las columnas por las que
        # el resumen agrupa, y asi el que reciba el archivo puede armar su
        # propia dinamica sin escribir una sola formula.
        if day:
            ws.cell(row=r, column=11, value=day.year)
            ws.cell(row=r, column=12, value=day.month)
            week = ws.cell(row=r, column=13, value=monday)
            week.number_format = _DATE_FORMAT

    last_row = header_row + len(movements)
    _as_table(ws, "TagInstalled", header_row, last_row, len(_TAG_COLUMNS))
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def _sheet_tag_summary(wb, rows: list, grain: str) -> None:
    """Resumen de instalacion por la granularidad elegida.

    Con grano mensual se conserva el par Ano | Mes del consolidado del cliente;
    con los otros granos una sola columna de periodo dice mas (un 'ano' repetido
    en 365 filas diarias no aporta nada).
    """
    ws = wb.create_sheet(i18n.t("sheet.tagsummary"))
    row = _title(ws, 1, "%s — %s" % (i18n.t("xls.title_tagsummary"),
                                     i18n.t("grain." + grain)),
                 i18n.t("xls.generated", when=_stamp()))
    header_row = row
    by_month = grain == analytics.GRAIN_MONTH
    labels = ([i18n.t("col.year"), i18n.t("col.month")] if by_month
              else [i18n.t("grain.col_" + grain)])
    labels += ["SMU", "TAG", i18n.t("col.total"),
               i18n.tr_value(tag_reader.MOVE_REMOVAL)]
    _headers(ws, header_row, labels,
             ([8, 10] if by_month else [14]) + [9, 9, 10, 12])

    first_data_col = 3 if by_month else 2
    for offset, item in enumerate(rows, start=1):
        r = header_row + offset
        day = item.date
        if by_month:
            ws.cell(row=r, column=1, value=day.year if day else None)
            ws.cell(row=r, column=2,
                    value=i18n.month_short(day.month) if day else item.period)
        elif grain == analytics.GRAIN_YEAR:
            ws.cell(row=r, column=1, value=day.year if day else item.period)
        else:
            # Dia y semana van como fecha real para que Excel las ordene y las
            # pueda agrupar en una dinamica propia.
            cell = ws.cell(row=r, column=1, value=day)
            cell.number_format = _DATE_FORMAT
        ws.cell(row=r, column=first_data_col, value=item.smu)
        ws.cell(row=r, column=first_data_col + 1, value=item.tag)
        ws.cell(row=r, column=first_data_col + 2,
                value=item.total).font = Font(bold=True)
        ws.cell(row=r, column=first_data_col + 3, value=item.removals)

    last_row = header_row + len(rows)
    if last_row <= header_row:
        return

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "%s — %s" % (i18n.t("xls.chart_taginstalled"),
                               i18n.t("grain." + grain))
    chart.y_axis.title = i18n.t("xls.axis_count")
    chart.x_axis.title = i18n.t("grain.col_" + grain)
    data = Reference(ws, min_col=first_data_col, max_col=first_data_col + 1,
                     min_row=header_row, max_row=last_row)
    label_col = 2 if by_month else 1
    cats = Reference(ws, min_col=label_col, max_col=label_col,
                     min_row=header_row + 1, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    _text_categories(chart, cats)
    chart.height, chart.width = 11.0, 26.0
    ws.add_chart(chart, "%s%d" % (get_column_letter(first_data_col + 5),
                                  header_row + 1))

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def export_tags(path: str, movements: list,
                grain: str = analytics.GRAIN_MONTH, progress_cb=None) -> str:
    """Escribe el consolidado de tags instalados y devuelve la ruta.

    `grain` agrupa el resumen por dia, semana, mes o ano. La tabla de
    movimientos no cambia: es el detalle, y ahi cada fila es un movimiento.
    """
    movements = list(movements or [])
    grain = grain if grain in analytics.GRAINS else analytics.GRAIN_MONTH
    summary = analytics.tag_by_period(movements, grain)

    wb = Workbook()
    wb.remove(wb.active)
    if progress_cb:
        progress_cb(1, 2, i18n.t("sheet.taginstalled"))
    _sheet_tag_data(wb, movements)
    if progress_cb:
        progress_cb(2, 2, i18n.t("sheet.tagsummary"))
    _sheet_tag_summary(wb, summary, grain)

    undated = sum(1 for m in movements if not m.get("date"))
    notes = [i18n.t("note.tags"), i18n.t("note.tags_device"),
             i18n.t("note.tags_dedupe")]
    if undated:
        notes.append(i18n.t("note.undated", n=undated))
    _sheet_notes(wb, notes)

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Nombres sugeridos
# ---------------------------------------------------------------------------
def suggested_name(kind: str) -> str:
    stamp = datetime.datetime.now().strftime("%y%m%d")
    if kind == "tags":
        base = ("Tags Instalados - Consolidado" if i18n.current() == i18n.ES
                else "Tag Installed - Consolidated")
    else:
        base = ("Mantenimiento de Flota - Reporte" if i18n.current() == i18n.ES
                else "Fleet Maintenance - Report")
    return "%s_%s.xlsx" % (base, stamp)


def default_path(directory: str, kind: str) -> str:
    return os.path.join(directory or os.path.expanduser("~"),
                        suggested_name(kind))
